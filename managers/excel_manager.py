from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from returns.result import Result, Success, Failure

def read_from_active_sheet(path_xlsx: str, *row_names)-> Result[dict, str]:
    ''' Return dictionary with keys as row names and values as list of table values'''
    header_row_num_10 = 10
    header_row_num_11 = 11

    try:
        wb = load_workbook(path_xlsx)
    except Exception as e:
        return Failure(f"Failed to open Excel file. {str(e)}")

    ws = wb.active

    header_10 = ws[header_row_num_10]
    header_11 = ws[header_row_num_11]

    '''# Get column letters as in Excel
    cols = [None for x in range(len(row_names))]
    for i, s in enumerate(row_names):
        col_letter = find_column_containing_str(header_10, s)
        cols[i] = col_letter'''

    '''# Get column values as list in dictionary
    col_dict_values = {}
    for row_name, col in zip(row_names, cols):
        if col != None:
            cells = ws[col][header_row_num_10:]
            col_dict_values[row_name] = [entry.value.strip() for entry in cells]'''

    cols_10 = get_column_letters(header_10, row_names)
    cols_11 = get_column_letters(header_11, row_names)

    col_dict_values_10 = get_column_values(ws, header_row_num_10, cols_10, row_names)
    col_dict_values_11 = get_column_values(ws, header_row_num_11, cols_11, row_names)

    col_dict_values = {**col_dict_values_10, **col_dict_values_11}      # Merge dicts and overwrite existing values(if exist)

    if len(set(map(len, col_dict_values.values())))!=1:
        return Failure(f"Failure. Columns in Excel document are of different length")

    return Success(col_dict_values)


def get_column_letters(header_num, row_names)-> list:
    '''Get column letters as in Excel (e.g. 'A' or 'BD')'''
    cols = [None for x in range(len(row_names))]
    for i, s in enumerate(row_names):
        col_letter = find_column_containing_str(header_num, s)
        cols[i] = col_letter
    return cols


def get_column_values(ws: Worksheet, header_num, cols, row_names)-> dict:
    '''Get column values as list in dictionary. Get cell values.'''
    col_dict_values = {}
    for row_name, col in zip(row_names, cols):
        if col != None:
            cells = ws[col][header_num:]
            col_dict_values[row_name] = [str(entry.value).strip() for entry in cells if entry.value != None]
    return col_dict_values


def find_column_containing_str(row, s: str):
    for cell in row:
        if cell.value is not None: 
            if s.lower() in cell.value.lower():
                return get_column_letter(cell.column)
    return None

def find_row_containing_str(col, s: str):
    for cell in col:
        if cell.value is not None: #We need to check that the cell is not empty.
            if s.lower() in cell.value.lower(): #Check if the value of the cell contains the string
                return cell.row
    return None